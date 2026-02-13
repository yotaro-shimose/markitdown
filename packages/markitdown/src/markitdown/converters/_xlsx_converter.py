import re
import sys
from pathlib import Path
from typing import Any, BinaryIO

import numpy as np
import numpy.typing as npt

# == Add by Masahiro (2026/01/28) ==
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from scipy.ndimage import label
from spire.xls import Workbook, XlsShape

from .._base_converter import DocumentConverter, DocumentConverterResult
from .._exceptions import MISSING_DEPENDENCY_MESSAGE, MissingDependencyException
from .._stream_info import StreamInfo
from ._html_converter import HtmlConverter

Mask2D = npt.NDArray[np.bool_]


def safe_name(name: str) -> str:
    # Windowsで使えない文字を置換
    return re.sub(r'[\\/:*?"<>|]', "_", name)


def _export_charts_via_excel_com(
    xlsx_path: Path, sheet_name: str, media_dir: Path, md_dir: Path
) -> list[str]:

    xlsx_path = Path(xlsx_path).resolve()
    media_dir = Path(media_dir).resolve()
    media_dir.mkdir(exist_ok=True)

    md_lines: list[str] = []

    try:
        workbook = Workbook()
        workbook.LoadFromFile(str(xlsx_path))
        for i in range(workbook.Worksheets.Count):
            worksheet = workbook.Worksheets.get_Item(i)
            ws_name = worksheet.Name  # ← いま処理中のシート名

            if ws_name != sheet_name:
                continue

            for j in range(worksheet.Charts.Count):
                print("チャート保存中")
                fname = f"{xlsx_path.stem}__{safe_name(ws_name)}__chart{j}.png"
                out_path = media_dir / fname

                chartImage = workbook.SaveChartAsImage(worksheet, j)
                chartImage.Save(str(out_path))

                md_lines.append(f"![]({media_dir.name}/{fname})")

                md_content = "### Charts\n"
                md_content += f"![]({media_dir.name}/{fname})\n\n"

                md_path = xlsx_path.with_name(
                    f"{xlsx_path.stem}__{safe_name(ws_name)}__chart{j}.md"
                )
                md_path.write_text(md_content, encoding="utf-8")

    except Exception as e:
        print(e)

    return md_lines


# チャートシートPNG化関数
def _export_chart_sheets_via_excel_com(
    xlsx_path: Path, sheet_name: str, media_dir: Path, md_dir: Path
) -> list[str]:

    xlsx_path = Path(xlsx_path).resolve()
    media_dir = Path(media_dir).resolve()
    media_dir.mkdir(exist_ok=True)

    md_lines: list[str] = []

    try:
        workbook = Workbook()
        workbook.LoadFromFile(str(xlsx_path))
        for i in range(workbook.Worksheets.Count):
            # ワークシートを取得
            worksheet = workbook.Worksheets.get_Item(i)
            ws_name = worksheet.Name

            if ws_name != sheet_name:
                continue

            # ワークシート内の図形を繰り返し処理
            for j in range(worksheet.PrstGeomShapes.Count):
                print("図形保存中")
                # 図形を取得
                shape = worksheet.PrstGeomShapes.get_Item(j)
                fname = f"{xlsx_path.stem}__{safe_name(ws_name)}__shape{j}.png"
                out_path = media_dir / fname

                # 図形をXlsShapeオブジェクトに変換
                xlsShape = XlsShape(shape)
                # 図形を画像ストリームとして保存
                image = xlsShape.SaveToImage()
                # 画像ストリームをファイルに保存
                image.Save(str(out_path))

                md_lines.append(f"![]({media_dir.name}/{fname})")

                md_content = "### Shapes\n"
                md_content += f"![]({media_dir.name}/{fname})\n\n"

                md_path = xlsx_path.with_name(
                    f"{xlsx_path.stem}__{safe_name(ws_name)}__shape{j}.md"
                )
                md_path.write_text(md_content, encoding="utf-8")

    except Exception as e:
        print(e)

    return md_lines


# Try loading optional (but in this case, required) dependencies
# Save reporting of any exceptions for later
_xlsx_dependency_exc_info = None
try:
    import openpyxl  # noqa: F401
    import pandas as pd
except ImportError:
    _xlsx_dependency_exc_info = sys.exc_info()

_xls_dependency_exc_info = None
try:
    import pandas as pd  # noqa: F811
    import xlrd  # noqa: F401
except ImportError:
    _xls_dependency_exc_info = sys.exc_info()

ACCEPTED_XLSX_MIME_TYPE_PREFIXES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
]
ACCEPTED_XLSX_FILE_EXTENSIONS = [".xlsx"]

ACCEPTED_XLS_MIME_TYPE_PREFIXES = [
    "application/vnd.ms-excel",
    "application/excel",
]
ACCEPTED_XLS_FILE_EXTENSIONS = [".xls"]


class XlsxConverter(DocumentConverter):
    """
    Converts XLSX files to Markdown, with each sheet presented as a separate Markdown table.
    """

    def __init__(self):
        super().__init__()
        self._html_converter = HtmlConverter()

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension in ACCEPTED_XLSX_FILE_EXTENSIONS:
            return True

        for prefix in ACCEPTED_XLSX_MIME_TYPE_PREFIXES:
            if mimetype.startswith(prefix):
                return True

        return False

    def count_nan_or_edge(
        self,
        mask: Mask2D,
        start_r: int,
        start_c: int,
        dr: int,
        dc: int,
    ) -> int:
        """
        mask から (start_r, start_c) の外側に向かって
        dr, dc 方向に何マス連続で
        - NaN
        - または端
        が続くかを数える
        """
        count = 0
        r, c = start_r + dr, start_c + dc

        while True:
            if r < 0 or r >= mask.shape[0] or c < 0 or c >= mask.shape[1]:
                count += 1
                break
            if mask[r, c]:
                break
            count += 1
            r += dr
            c += dc

        return count

    # 周囲が NaN（または端）かチェック
    def is_nan_or_edge(self, mask: Mask2D, r: int, c: int) -> bool:
        if r < 0 or r >= mask.shape[0]:
            return True
        if c < 0 or c >= mask.shape[1]:
            return True
        return not bool(mask[r, c])

    def proc_md_content(
        self,
        s: str,
        sheets: dict[str, pd.DataFrame],
        wb_img: OpenpyxlWorkbook,
        xlsx_path: Path,
        media_dir: Path,
        md_dir: Path,
        **kwargs: Any,
    ) -> str:

        df = sheets[s].replace(
            r"^\s*$", pd.NA, regex=True
        )  # 空文字も空扱いにする（不要なら削除OK）

        ## 矩形判定
        mask = df.notna().to_numpy()

        # --- 1列のNaN（T F T）を埋める ---
        for c in range(1, mask.shape[1] - 1):
            mask[:, c] |= mask[:, c - 1] & mask[:, c + 1]

        # --- 2列のNaN（T F F T）を埋める ---
        for c in range(1, mask.shape[1] - 2):
            cond = ~mask[:, c] & ~mask[:, c + 1] & mask[:, c - 1] & mask[:, c + 2]
            mask[cond, c] = True
            mask[cond, c + 1] = True

        # 1行NaN
        for r in range(1, mask.shape[0] - 1):
            mask[r, :] |= mask[r - 1, :] & mask[r + 1, :]

        # 2行NaN
        for r in range(1, mask.shape[0] - 2):
            cond = ~mask[r, :] & ~mask[r + 1, :] & mask[r - 1, :] & mask[r + 2, :]
            mask[r, cond] = True
            mask[r + 1, cond] = True

        # 4近傍で連結成分を抽出
        # label: Trueが隣接している塊に同じ番号を振る関数
        labeled, num = label(mask)

        blocks = []

        for lab in range(1, num + 1):
            coords = np.argwhere(labeled == lab)
            r0, c0 = coords.min(axis=0)
            r1, c1 = coords.max(axis=0)

            # 各方向の NaN 幅を計算
            top_gap = min(
                self.count_nan_or_edge(mask, r0, c, -1, 0) for c in range(c0, c1 + 1)
            )
            bottom_gap = min(
                self.count_nan_or_edge(mask, r1, c, 1, 0) for c in range(c0, c1 + 1)
            )
            left_gap = min(
                self.count_nan_or_edge(mask, r, c0, 0, -1) for r in range(r0, r1 + 1)
            )
            right_gap = min(
                self.count_nan_or_edge(mask, r, c1, 0, 1) for r in range(r0, r1 + 1)
            )

            nan_width_ok = (
                top_gap >= 2 or bottom_gap >= 2 or left_gap >= 2 or right_gap >= 2
            )

            if not nan_width_ok:
                continue

            top_ok = all(
                self.is_nan_or_edge(mask, r0 - 1, c) for c in range(c0, c1 + 1)
            )
            bottom_ok = all(
                self.is_nan_or_edge(mask, r1 + 1, c) for c in range(c0, c1 + 1)
            )
            left_ok = all(
                self.is_nan_or_edge(mask, r, c0 - 1) for r in range(r0, r1 + 1)
            )
            right_ok = all(
                self.is_nan_or_edge(mask, r, c1 + 1) for r in range(r0, r1 + 1)
            )

            if not (top_ok and bottom_ok and left_ok and right_ok):
                continue

            blocks.append((r0, r1, c0, c1))

        for r0, r1, c0, c1 in blocks:
            md_content = f"## {s}_{r0}_{r1}_{c0}_{c1}\n"
            md_name = f"{s}_{r0}_{r1}_{c0}_{c1}"
            extracted = df.iloc[r0 : r1 + 1, c0 : c1 + 1]
            html_content = extracted.to_html(index=False)

            md_content += (
                self._html_converter.convert_string(
                    html_content, **kwargs
                ).markdown.strip()
                + "\n\n"
            )

            md_path = xlsx_path.with_name(
                f"{xlsx_path.stem}_{safe_name(s)}_{md_name}.md"
            )
            print("md_path:", md_path)
            md_path.write_text(md_content, encoding="utf-8")

        ws = wb_img[s]

        # シート内画像の保存
        images = getattr(ws, "_images", [])
        if images:
            for i, img in enumerate(images, start=1):
                try:
                    data = img._data()
                except Exception:
                    continue

                md_content = "### Images\n"

                fname = f"{xlsx_path.stem}__{s}__img{i}.png"
                (media_dir / fname).write_bytes(data)
                md_content += f"![]({media_dir.name}/{fname})\n\n"

                md_path = xlsx_path.with_name(f"{xlsx_path.stem}__{s}__img{i}.md")
                md_path.write_text(md_content, encoding="utf-8")

        _ = _export_charts_via_excel_com(xlsx_path, s, media_dir, md_dir)

        return md_content

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> DocumentConverterResult:

        # Check the dependencies
        if _xlsx_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xlsx",
                    feature="xlsx",
                )
            ) from _xlsx_dependency_exc_info[1].with_traceback(  # type: ignore[union-attr]
                _xlsx_dependency_exc_info[2]
            )

        sheets = pd.read_excel(file_stream, sheet_name=None, engine="openpyxl")

        if stream_info.local_path is None:
            raise ValueError("stream_info.local_path is None")

        xlsx_path = Path(stream_info.local_path).resolve()
        media_dir = xlsx_path.parent / "media"
        media_dir.mkdir(exist_ok=True)

        md_dir = xlsx_path.parent / "md"
        md_dir.mkdir(exist_ok=True)

        file_stream.seek(0)
        wb_img = load_workbook(file_stream, data_only=True)

        md_content = ""
        for sheet_idx, s in enumerate(sheets, start=1):
            print(s)
            md_content = self.proc_md_content(
                s, sheets, wb_img, xlsx_path, media_dir, md_dir, **kwargs
            )

        return DocumentConverterResult(markdown=md_content.strip())


class XlsConverter(DocumentConverter):
    """
    Converts XLS files to Markdown, with each sheet presented as a separate Markdown table.
    """

    def __init__(self):
        super().__init__()
        self._html_converter = HtmlConverter()

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension in ACCEPTED_XLS_FILE_EXTENSIONS:
            return True

        for prefix in ACCEPTED_XLS_MIME_TYPE_PREFIXES:
            if mimetype.startswith(prefix):
                return True

        return False

    def count_nan_or_edge(
        self,
        mask: Mask2D,
        start_r: int,
        start_c: int,
        dr: int,
        dc: int,
    ) -> int:
        """
        mask から (start_r, start_c) の外側に向かって
        dr, dc 方向に何マス連続で
        - NaN
        - または端
        が続くかを数える
        """
        count = 0
        r, c = start_r + dr, start_c + dc

        while True:
            if r < 0 or r >= mask.shape[0] or c < 0 or c >= mask.shape[1]:
                count += 1
                break
            if mask[r, c]:
                break
            count += 1
            r += dr
            c += dc

        return count

    # 周囲が NaN（または端）かチェック
    def is_nan_or_edge(self, mask: Mask2D, r: int, c: int) -> bool:
        if r < 0 or r >= mask.shape[0]:
            return True
        if c < 0 or c >= mask.shape[1]:
            return True
        return not bool(mask[r, c])

    def proc_md_content(
        self,
        s: str,
        sheets: dict[str, pd.DataFrame],
        wb_img: OpenpyxlWorkbook,
        xlsx_path: Path,
        media_dir: Path,
        md_dir: Path,
        **kwargs: Any,
    ) -> str:

        df = sheets[s].replace(
            r"^\s*$", pd.NA, regex=True
        )  # 空文字も空扱いにする（不要なら削除OK）

        ## 矩形判定
        mask = df.notna().to_numpy()

        # --- 1列のNaN（T F T）を埋める ---
        for c in range(1, mask.shape[1] - 1):
            mask[:, c] |= mask[:, c - 1] & mask[:, c + 1]

        # --- 2列のNaN（T F F T）を埋める ---
        for c in range(1, mask.shape[1] - 2):
            cond = ~mask[:, c] & ~mask[:, c + 1] & mask[:, c - 1] & mask[:, c + 2]
            mask[cond, c] = True
            mask[cond, c + 1] = True

        # 1行NaN
        for r in range(1, mask.shape[0] - 1):
            mask[r, :] |= mask[r - 1, :] & mask[r + 1, :]

        # 2行NaN
        for r in range(1, mask.shape[0] - 2):
            cond = ~mask[r, :] & ~mask[r + 1, :] & mask[r - 1, :] & mask[r + 2, :]
            mask[r, cond] = True
            mask[r + 1, cond] = True

        # 4近傍で連結成分を抽出
        # label: Trueが隣接している塊に同じ番号を振る関数
        labeled, num = label(mask)

        blocks = []

        for lab in range(1, num + 1):
            coords = np.argwhere(labeled == lab)
            r0, c0 = coords.min(axis=0)
            r1, c1 = coords.max(axis=0)

            # 各方向の NaN 幅を計算
            top_gap = min(
                self.count_nan_or_edge(mask, r0, c, -1, 0) for c in range(c0, c1 + 1)
            )
            bottom_gap = min(
                self.count_nan_or_edge(mask, r1, c, 1, 0) for c in range(c0, c1 + 1)
            )
            left_gap = min(
                self.count_nan_or_edge(mask, r, c0, 0, -1) for r in range(r0, r1 + 1)
            )
            right_gap = min(
                self.count_nan_or_edge(mask, r, c1, 0, 1) for r in range(r0, r1 + 1)
            )

            nan_width_ok = (
                top_gap >= 2 or bottom_gap >= 2 or left_gap >= 2 or right_gap >= 2
            )

            if not nan_width_ok:
                continue

            top_ok = all(
                self.is_nan_or_edge(mask, r0 - 1, c) for c in range(c0, c1 + 1)
            )
            bottom_ok = all(
                self.is_nan_or_edge(mask, r1 + 1, c) for c in range(c0, c1 + 1)
            )
            left_ok = all(
                self.is_nan_or_edge(mask, r, c0 - 1) for r in range(r0, r1 + 1)
            )
            right_ok = all(
                self.is_nan_or_edge(mask, r, c1 + 1) for r in range(r0, r1 + 1)
            )

            if not (top_ok and bottom_ok and left_ok and right_ok):
                continue

            blocks.append((r0, r1, c0, c1))

        for r0, r1, c0, c1 in blocks:
            md_content = f"## {s}_{r0}_{r1}_{c0}_{c1}\n"
            md_name = f"{s}_{r0}_{r1}_{c0}_{c1}"
            extracted = df.iloc[r0 : r1 + 1, c0 : c1 + 1]
            html_content = extracted.to_html(index=False)
            md_content += (
                self._html_converter.convert_string(
                    html_content, **kwargs
                ).markdown.strip()
                + "\n\n"
            )

            md_path = xlsx_path.with_name(
                f"{xlsx_path.stem}_{safe_name(s)}_{md_name}.md"
            )
            print("md_path:", md_path)
            md_path.write_text(md_content, encoding="utf-8")

        ws = wb_img[s]

        # シート内画像の保存
        images = getattr(ws, "_images", [])
        if images:
            for i, img in enumerate(images, start=1):
                try:
                    data = img._data()
                except Exception:
                    continue

                md_content = "### Images\n"

                fname = f"{xlsx_path.stem}__{s}__img{i}.png"
                (media_dir / fname).write_bytes(data)
                md_content += f"![]({media_dir.name}/{fname})\n\n"

                md_path = xlsx_path.with_name(f"{xlsx_path.stem}__{s}__img{i}.md")
                md_path.write_text(md_content, encoding="utf-8")

        _ = _export_charts_via_excel_com(xlsx_path, s, media_dir, md_dir)

        return md_content

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> DocumentConverterResult:

        # Check the dependencies
        if _xlsx_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xlsx",
                    feature="xlsx",
                )
            ) from _xlsx_dependency_exc_info[1].with_traceback(  # type: ignore[union-attr]
                _xlsx_dependency_exc_info[2]
            )

        sheets = pd.read_excel(file_stream, sheet_name=None, engine="openpyxl")

        if stream_info.local_path is None:
            raise ValueError("stream_info.local_path is None")

        xlsx_path = Path(stream_info.local_path).resolve()
        media_dir = xlsx_path.parent / "media"
        media_dir.mkdir(exist_ok=True)

        md_dir = xlsx_path.parent / "md"
        md_dir.mkdir(exist_ok=True)

        file_stream.seek(0)
        wb_img = load_workbook(file_stream, data_only=True)

        md_content = ""
        for sheet_idx, s in enumerate(sheets, start=1):
            print(s)
            md_content = self.proc_md_content(
                s, sheets, wb_img, xlsx_path, media_dir, md_dir, **kwargs
            )

        return DocumentConverterResult(markdown=md_content.strip())
